# -*- coding: utf-8 -*-
# Copyright (c) 2012-2013 by Pablo Martín <goinnn@gmail.com>
#
# This software is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This software is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this software.  If not, see <http://www.gnu.org/licenses/>.

import csv
import collections
import sys

import openpyxl

from openpyxl import Workbook
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from .base import get_content

PY3 = sys.version_info[0] == 3

OPENPYXL_VERSION = tuple([int(x) for x in openpyxl.__version__.split(".")]) # (openpyxl.__major__, openpyxl.__minor__, openpyxl.__release__)
INITIAL = 0

if OPENPYXL_VERSION >= (2,0,0):
    INITIAL = 1

def convert(response, encoding='utf-8', title_sheet='Sheet 1', content_attr='content', csv_kwargs=None):
    csv_kwargs = csv_kwargs or {}
    try:
        wb = Workbook(encoding=encoding)
    except TypeError:
        wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = title_sheet
    cell_widths = collections.defaultdict(lambda: 0)
    content = get_content(response, encoding=encoding, content_attr=content_attr)
    reader = csv.reader(content, **csv_kwargs)

    for lno, line in enumerate(reader, INITIAL):
        write_row(ws, lno, line, cell_widths, encoding=encoding)

    # Roughly autosize output column widths based on maximum column size
    # and add bold style for the header
    for i, cell_width in cell_widths.items():
        cell = ws.cell(column=i, row=INITIAL)
        if OPENPYXL_VERSION>=(2,0,0):
            bold = cell.font.copy(bold=True)
            cell.font = bold
        else:
            cell.style.font.bold = True
        ws.column_dimensions[get_column_letter(i + 1)].width = cell_width

    setattr(response, content_attr, '')
    wb.save(response)


def write_row(ws, lno, cell_text, cell_widths, encoding='utf-8'):
    for cno, cell_text in enumerate(cell_text, INITIAL):
        if not PY3:
            cell_text = cell_text.decode(encoding)
        ws.cell(column=cno, row=lno).value = cell_text
        cell_widths[cno] = max(
            cell_widths[cno],
            len(cell_text))
